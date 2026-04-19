"""
Convert validation_log.csv into a formatted Excel workbook with summary
sheet. Handles the capstone's live-capture CSV template even before rows
are filled in.

Usage:
    cd final_capstone
    python3 extras/build_validation_xlsx.py

Output: extras/STAR_ValidationLog.xlsx
"""

import csv
from pathlib import Path
from collections import defaultdict
import statistics

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
SRC = HERE / "validation_log.csv"
OUT = HERE / "STAR_ValidationLog.xlsx"

MAROON = "500000"
LIGHT = "F5F5F7"


def main():
    wb = Workbook()
    raw = wb.active
    raw.title = "Raw measurements"

    rows = []
    with SRC.open() as f:
        for row in csv.reader(f):
            if not row:
                continue
            first = row[0].strip() if row else ""
            if first.startswith("#"):
                continue
            rows.append(row)

    if not rows:
        raise RuntimeError(f"validation_log.csv at {SRC} is empty")

    header = rows[0]
    width = len(header)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor=MAROON)
    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row in enumerate(rows, 1):
        padded = row + [""] * (width - len(row))
        for c, val in enumerate(padded[:width], 1):
            cell = raw.cell(row=r, column=c, value=val)
            cell.border = border
            if r == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col in range(1, width + 1):
        letter = get_column_letter(col)
        try:
            max_len = max(
                len(str((r + [""] * (width - len(r)))[col - 1]))
                for r in rows
            )
        except ValueError:
            max_len = 12
        raw.column_dimensions[letter].width = min(max_len + 2, 32)

    summary = wb.create_sheet("Summary by check")
    sum_header = ["Check type", "n accepted",
                  "Mean slope (deg)", "Stdev (deg)", "Max error (deg)",
                  "ADA section"]
    for c, h in enumerate(sum_header, 1):
        cell = summary.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Build summary from the data rows (skip templates with empty values)
    idx = {h: i for i, h in enumerate(header)}
    data_rows = rows[1:]

    errs = []
    for r in data_rows:
        padded = r + [""] * (width - len(r))
        try:
            gt = float(padded[idx["ramp_slope_gt_deg"]])
            lidar = float(padded[idx["ramp_slope_lidar_deg"]])
        except (KeyError, ValueError):
            continue
        errs.append(abs(gt - lidar))

    if errs:
        mean = round(statistics.mean(errs), 3)
        sd = round(statistics.stdev(errs), 3) if len(errs) > 1 else 0.0
        mx = round(max(errs), 3)
    else:
        mean = "TBD"
        sd = "TBD"
        mx = "TBD"

    row_values = ["Ramp slope (405.2)", len(errs), mean, sd, mx, "ADA 405.2"]
    for c, v in enumerate(row_values, 1):
        cell = summary.cell(row=2, column=c, value=v)
        cell.border = border

    for col in range(1, len(sum_header) + 1):
        summary.column_dimensions[get_column_letter(col)].width = 20

    # Notes sheet
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Ground truth: Wixey WR300 digital angle gauge (+/- 0.1 deg)."
    notes["A2"] = "Cross-validation gate between BNO055 pitch and LiDAR plane normal: +/- 0.5 deg."
    notes["A3"] = "ADA section 405.2 ramp slope threshold: 4.76 degrees."
    notes["A4"] = "Fill validation_log.csv with one row per ramp measurement on the on-campus ramp."
    notes["A5"] = "Re-run build_validation_xlsx.py after each capture to refresh this workbook."

    wb.save(str(OUT))
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
